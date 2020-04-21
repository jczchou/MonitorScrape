# -*- coding: utf-8 -*-
"""
Created on 3-27-2020   @author: jchou
AIM: DOWNLOAD ALL THE MONITOR & TV SPECS ON www.displayspecifications.com
"""
import time
import openpyxl
import bs4 as bs
import requests
from urlextract import URLExtract
from openpyxl.utils import get_column_letter
import re

# get_tor_session uses tor to spoof IP address, each time a request is made 
# to scrape a page, make sure to launch a new instance of session by calling 
# get_tor_session to avoid being blocked by the scraped site
def get_tor_session():  
    session = requests.session()
    # uses the 9150 port though 9050 is the default TOR socks port
    session.proxies = {'http':  'socks5://127.0.0.1:9150',
                       'https': 'socks5://127.0.0.1:9150'}
    return session

def OEM_list_initialize():
    OEM_list = []  # list to store dictionary variables of each OEM Brand & URL
    
    Acer_dict = {'brand': 'Acer', 'url': 'https://www.displayspecifications.com/en/brand/f652d'}      
    AOC_dict = {'brand': 'AOC', 'url': 'https://www.displayspecifications.com/en/brand/1fe710'}
    AOpen_dict = {'brand': 'AOpen', 'url': 'https://www.displayspecifications.com/en/brand/98142e'}        
    Apple_dict = {'brand': 'Apple', 'url': 'https://www.displayspecifications.com/en/brand/368c36'}
    Coolermaster_dict = {'brand': 'Coolermaster', 'url': 'https://www.displayspecifications.com/en/brand/505a43'}        
    ASUS_dict = {'brand': 'ASUS', 'url': 'https://www.displayspecifications.com/en/brand/db0f8'}
    BenQ_dict = {'brand': 'BenQ', 'url': 'https://www.displayspecifications.com/en/brand/62dea'}
    Coolermaster_dict = {'brand': 'Coolermaster', 'url': 'https://www.displayspecifications.com/en/brand/505a43'}
    Dell_dict = {'brand': 'Dell', 'url': 'https://www.displayspecifications.com/en/brand/91913'}
    EIZO_dict = {'brand': 'EIZO', 'url': 'https://www.displayspecifications.com/en/brand/04692'}
    Fujitsu_dict = {'brand': 'Fujitsu', 'url': 'https://www.displayspecifications.com/en/brand/a7df1f'}
    Gamemax_dict = {'brand': 'Gamemax', 'url': 'https://www.displayspecifications.com/en/brand/a47f2b'}    
    Gigabyte_dict = {'brand': 'Gigabyte', 'url': 'https://www.displayspecifications.com/en/brand/669331'}
    Hannspree_dict= {'brand': 'Hannspree', 'url': 'https://www.displayspecifications.com/en/brand/e50329'}
    HP_dict = {'brand': 'HP', 'url': 'https://www.displayspecifications.com/en/brand/49fd19'}    
    Iiyama_dict = {'brand': 'Iiyama', 'url': 'https://www.displayspecifications.com/en/brand/cb741b'}
    Lenovo_dict = {'brand': 'Lenovo', 'url': 'https://www.displayspecifications.com/en/brand/cd481e'}
    LG_dict = {'brand': 'LG', 'url': 'https://www.displayspecifications.com/en/brand/a1025'}
    Medion_dict = {'brand': 'Medion', 'url': 'https://www.displayspecifications.com/en/brand/fdfa2a'}
    MSI_dict = {'brand': 'MSI', 'url': 'https://www.displayspecifications.com/en/brand/fa4f21'}
    NEC_dict = {'brand': 'NEC', 'url': 'https://www.displayspecifications.com/en/brand/c5e89'}
    Philips_dict = {'brand': 'Philips', 'url': 'https://www.displayspecifications.com/en/brand/3750b'}
    Pixio_dict = {'brand': 'Pixio', 'url': 'https://www.displayspecifications.com/en/brand/e19f3d'}
    QNIX_dict = {'brand': 'QNIX', 'url': 'https://www.displayspecifications.com/en/brand/b69c37'}
    Razer_dict = {'brand': 'Razer', 'url': 'https://www.displayspecifications.com/en/brand/536c30'}
    Samsung_dict = {'brand': 'Samsung', 'url': 'https://www.displayspecifications.com/en/brand/08cd6'}
    Sharp_dict = {'brand': 'Sharp', 'url': 'https://www.displayspecifications.com/en/brand/869211'}
    ViewSonic_dict = {'brand': 'ViewSonic', 'url': 'https://www.displayspecifications.com/en/brand/b65513'}
    Xiaomi_dict = {'brand': 'Xiaomi', 'url': 'https://www.displayspecifications.com/en/brand/6a897'}
    Yashi_dict = {'brand': 'Yashi', 'url': 'https://www.displayspecifications.com/en/brand/061033'}
    
    #OEM_list = [Acer_dict, AOC_dict, AOpen_dict, Apple_dict, ASUS_dict, BenQ_dict, Coolermaster_dict, Dell_dict, EIZO_dict, Fujitsu_dict, Gamemax_dict, Gigabyte_dict, Hannspree_dict, HP_dict, Iiyama_dict, Lenovo_dict, LG_dict, Medion_dict, MSI_dict, NEC_dict, Philips_dict, Pixio_dict, QNIX_dict, Razer_dict, Samsung_dict, Sharp_dict, ViewSonic_dict, Xiaomi_dict, Yashi_dict]

    OEM_list = [LG_dict]
    return OEM_list

    
# get the URL and return the scraped HTML
# for now, the scraping happens one vendor at a time. Ideally the URLs below 
# should be in an iterable list one winds thru to scrape but recent web site
# blocking efforts made this a bit tricky
def main(session_var, each_OEM):
    # use session instead of requests to avoid being blocked    
    rawHTML = session.get(str(each_OEM), headers={'User-Agent': 'Mozilla/5.0'}).text    
    return rawHTML


def find_models(raw):
    model_dict_list = []
    
    soup = bs.BeautifulSoup(raw, 'lxml')
    # grabs the model name into a list on the front page
    # year launched elem is the year a model launched    
    year_launched = soup.find_all("h1", class_="header")

    # Only find the DIV tag that has the models we want, not the "Last
    # View models" that has other OEM SKUs
    model_names = soup.find_all('div', class_='model-listing-container-80')        
    
    n=0  # set iterator counter to reference the year_launched list elements    
    # for each <DIV></DIV> section, find just the models in that section, such
    # as only 2020 monitors in one section
    for each_model in model_names:
        model_names3 = bs.BeautifulSoup(str(each_model), 'lxml')
        model_names4 = model_names3.find_all(['h3', 'a href'])
     
        for model_names5 in model_names4:
            model_dict = {'year': str(year_launched[n].text), 'model_url': str(model_names5)}
            model_dict_list.append(model_dict)
        n += 1        
    
    return(model_dict_list)


def parse_models_href(model_names):
    urls = [] # get cleaned list of URLs for all models of a brand
    
    for elem in model_names:
        url = extractor.find_urls(str(elem["model_url"]))
        url2 = {'year': elem["year"], 'model_url': str(url)}
        urls.append(url2)        

    urls = list(urls)
    return urls


def goto_url(urls):
    each_url = urls  #each_url = urls as code below already use 'each_url'
    output_rows = []  #a list to store each row of the table
    k=1  # row counter to keep track of rows
    
    # for each url get the specs
    for i in range(len(each_url)):
        # get rid of [, ], ' around the URL
        each_url[i]["model_url"] = each_url[i]["model_url"].replace("[", "")
        each_url[i]["model_url"] = each_url[i]["model_url"].replace("]", "")
        each_url[i]["model_url"] = each_url[i]["model_url"].replace("'", "")        
        each_url_cleaned = str(each_url[i]["model_url"])
        
        spec_for_ea_url = session.get(each_url_cleaned, headers={'User-Agent': 'Mozilla/5.0'}).text
        time.sleep(2)  #wait 3 secs to further avoid site scraping suspicion
        soup2 = bs.BeautifulSoup(spec_for_ea_url, 'lxml')
        
        tables = soup2.find_all("table", class_="model-information-table row-selection")
        
        # PATTERNS TO LOOK FOR ALL TEXT WITHIN VARIOUS TAGS
        p_pattern = '<p>(.+)</p>' # look for text with <p> & </p>
        br_pattern = '<br/>(.+)</td>' # look for text with <br> & </br>
        span_pattern = '<span(.*)></span>' # look for text with <span class=> & </span>        
        span_pattern2 = '<span>(.*)</span>' # further look for text with only <span> & </span>                
        span_pattern3 = '<span class\=\"arrow-bullet\"></span>'
        replace = '' # replace with blank if found

        # first insert the Year of the model b4 specs
        year_make_location = "A" + str(k)  #cell A1 for 1st model
        year_make_location2 = "B" + str(k)        
        sheet[year_make_location].value = "Year"
        sheet[year_make_location2].value = str(each_url[i]["year"])            
        k += 1

        for table in tables:
            # table is one single table for one SKU which like Brand section
            # iterate thru rach table row <tr> of specs
            
            for table_row in table.findAll('tr'):                
                # remove all text between <p>, <br>, <span> tags using pattern variables above
                table_row = re.sub(p_pattern, replace, str(table_row))

                # Connectivity field on site: we need to capture
                # all the different ports which is delineated by <span> tags
                # substitute <br/> with <td></td> to separate different ports
                # into distinct HTML table cells                
                Connectivity_filter = re.search('Connectivity', table_row)
                if Connectivity_filter:
                    table_row = re.sub(span_pattern3, replace, str(table_row))
                    table_row = re.sub('<br/>', '</td><td>', str(table_row))
                else:
                    table_row = re.sub(br_pattern, replace, str(table_row))                  
                    table_row = re.sub(span_pattern, replace, str(table_row))                
                    table_row = re.sub(span_pattern2, replace, str(table_row))

                # turn each table_row back into a Beautifulsoup object so
                # we can use find_call()
                soup3 = bs.BeautifulSoup(table_row, 'html.parser') 

                columns = soup3.find_all('td')
                output_td = []  # a list to store each cell of one row
                
                # iterate thru each <td> cell within a row
                j = 1
                
                for column in columns:
                    input_cell = get_column_letter(j) + str(k)
                    sheet[input_cell].value = str(column.text)
                    # append all the <td> elements in one <tr> row
                    output_td.append(column.text)
                    
                    j += 1
                    if str(j) == str(len(columns)):
                        output_rows.append(output_td)
                
                k += 1  # move to the next row for the next spec
        k += 1  # this inserts one blank row between models          
        output_rows.append('')
        
    return(output_rows)
    

if __name__ == '__main__':    
    #open the workbook used to store the specs
    wb = openpyxl.load_workbook('DisplaySpecifications_Scrape4.xlsx')    
    OEM_Master_List = OEM_list_initialize()  # get list of all OEMs to scrape
    ses = get_tor_session()
    masterurl_List = []

    for each_OEM in OEM_Master_List:
        sheet = wb.create_sheet(str(each_OEM["brand"]))  # worksheet for ea OEM
        raw = main(ses, each_OEM["url"])
        extractor = URLExtract()
        mn_list = find_models(raw)
    
    cleaned_url = parse_models_href(mn_list)
    
    if not masterurl_List:
        masterurl_List = cleaned_url
    else:
        masterurl_List = set(cleaned_url) - set(masterurl_List)
        masterurl_List = set(masterurl_List)
        masterurl_List = list(masterurl_List) 
    
    spec_tables = goto_url(masterurl_List)
    wb.save('DisplaySpecifications_Scrape4.xlsx')